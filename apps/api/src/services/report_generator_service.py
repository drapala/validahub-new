"""
Report generation service for multiple formats (Excel, PDF, JSON).
"""

import io
import json
from datetime import datetime
from typing import Dict, Any, Optional, List
import pandas as pd

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from ..core.logging_config import get_logger

logger = get_logger(__name__)


class ReportGeneratorService:
    """Service for generating validation reports in multiple formats."""
    
    def __init__(self):
        """Initialize the report generator service."""
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom PDF styles."""
        # Title style
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=TA_CENTER
        ))
        
        # Subtitle style
        self.styles.add(ParagraphStyle(
            name='CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#333333'),
            spaceAfter=20
        ))
    
    def generate_report(self, 
                       validation_result: Dict[str, Any], 
                       format: str = 'json',
                       filename: Optional[str] = None) -> bytes:
        """
        Generate validation report in specified format.
        
        Args:
            validation_result: The validation result dictionary
            format: Output format ('json', 'excel', 'pdf')
            filename: Optional filename for the report
            
        Returns:
            Report as bytes
        """
        format = format.lower()
        
        if format == 'json':
            return self._generate_json_report(validation_result)
        elif format == 'excel' or format == 'xlsx':
            return self._generate_excel_report(validation_result, filename)
        elif format == 'pdf':
            return self._generate_pdf_report(validation_result, filename)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def _generate_json_report(self, validation_result: Dict[str, Any]) -> bytes:
        """Generate JSON report."""
        return json.dumps(validation_result, indent=2, ensure_ascii=False).encode('utf-8')
    
    def _generate_excel_report(self, validation_result: Dict[str, Any], filename: Optional[str] = None) -> bytes:
        """Generate Excel report with multiple sheets."""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: Summary
            summary_data = {
                'Métrica': [
                    'Total de Linhas',
                    'Erros Críticos',
                    'Avisos',
                    'Total de Problemas',
                    'Tempo de Processamento (s)',
                    'Marketplace',
                    'Categoria'
                ],
                'Valor': [
                    validation_result['summary']['total_rows'],
                    validation_result['summary']['errors'],
                    validation_result['summary']['warnings'],
                    validation_result['summary']['total_issues'],
                    f"{validation_result['summary']['processing_time']:.4f}",
                    validation_result.get('marketplace', 'N/A'),
                    validation_result.get('category', 'N/A')
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Resumo', index=False)
            
            # Sheet 2: Issues Details
            if validation_result.get('issues'):
                issues_data = []
                for issue in validation_result['issues']:
                    issues_data.append({
                        'Linha': issue['row'],
                        'Severidade': issue['severity'],
                        'Campo': issue['field'],
                        'Problema': issue['message'],
                        'Valor Atual': issue.get('current_value', ''),
                        'Valor Esperado': issue.get('expected', '')
                    })
                df_issues = pd.DataFrame(issues_data)
                df_issues.to_excel(writer, sheet_name='Problemas Detalhados', index=False)
            
            # Sheet 3: Statistics by Field
            if validation_result.get('issues'):
                field_stats = {}
                for issue in validation_result['issues']:
                    field = issue['field']
                    severity = issue['severity']
                    
                    if field not in field_stats:
                        field_stats[field] = {'ERROR': 0, 'WARNING': 0}
                    field_stats[field][severity] += 1
                
                stats_data = []
                for field, counts in field_stats.items():
                    stats_data.append({
                        'Campo': field,
                        'Erros': counts['ERROR'],
                        'Avisos': counts['WARNING'],
                        'Total': counts['ERROR'] + counts['WARNING']
                    })
                df_stats = pd.DataFrame(stats_data)
                df_stats = df_stats.sort_values('Total', ascending=False)
                df_stats.to_excel(writer, sheet_name='Estatísticas por Campo', index=False)
            
            # Format the Excel file
            workbook = writer.book
            
            # Format Summary sheet
            summary_sheet = workbook['Resumo']
            self._format_excel_sheet(summary_sheet, is_summary=True)
            
            # Format Issues sheet
            if 'Problemas Detalhados' in workbook.sheetnames:
                issues_sheet = workbook['Problemas Detalhados']
                self._format_excel_sheet(issues_sheet, is_issues=True)
            
            # Format Statistics sheet
            if 'Estatísticas por Campo' in workbook.sheetnames:
                stats_sheet = workbook['Estatísticas por Campo']
                self._format_excel_sheet(stats_sheet)
        
        output.seek(0)
        return output.read()
    
    def _format_excel_sheet(self, worksheet, is_summary=False, is_issues=False):
        """Apply formatting to Excel worksheet."""
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Specific formatting for issues sheet
        if is_issues:
            for row in worksheet.iter_rows(min_row=2):
                severity_cell = row[1]  # Severidade column
                if severity_cell.value == 'ERROR':
                    severity_cell.font = Font(color="FF0000", bold=True)
                elif severity_cell.value == 'WARNING':
                    severity_cell.font = Font(color="FFA500", bold=True)
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    def _generate_pdf_report(self, validation_result: Dict[str, Any], filename: Optional[str] = None) -> bytes:
        """Generate PDF report."""
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        story = []
        
        # Title
        title = Paragraph(
            f"Relatório de Validação - {validation_result.get('marketplace', 'N/A')}",
            self.styles['CustomTitle']
        )
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Timestamp
        timestamp = datetime.fromisoformat(validation_result['timestamp'].replace('Z', '+00:00'))
        date_str = timestamp.strftime("%d/%m/%Y %H:%M:%S")
        story.append(Paragraph(f"Data: {date_str}", self.styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Summary Section
        story.append(Paragraph("Resumo da Validação", self.styles['CustomSubtitle']))
        
        summary_data = [
            ['Métrica', 'Valor'],
            ['Total de Linhas', str(validation_result['summary']['total_rows'])],
            ['Erros Críticos', str(validation_result['summary']['errors'])],
            ['Avisos', str(validation_result['summary']['warnings'])],
            ['Total de Problemas', str(validation_result['summary']['total_issues'])],
            ['Tempo de Processamento', f"{validation_result['summary']['processing_time']:.4f}s"],
            ['Marketplace', validation_result.get('marketplace', 'N/A')],
            ['Categoria', validation_result.get('category', 'N/A')]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # Issues Section
        if validation_result.get('issues'):
            story.append(Paragraph("Problemas Encontrados", self.styles['CustomSubtitle']))
            
            # Split issues into errors and warnings
            errors = [i for i in validation_result['issues'] if i['severity'] == 'ERROR']
            warnings = [i for i in validation_result['issues'] if i['severity'] == 'WARNING']
            
            # Errors table
            if errors:
                story.append(Paragraph("Erros Críticos", self.styles['Heading3']))
                error_data = [['Linha', 'Campo', 'Problema', 'Valor Atual']]
                
                for error in errors[:20]:  # Limit to 20 for PDF
                    error_data.append([
                        str(error['row']),
                        error['field'],
                        error['message'][:50] + '...' if len(error['message']) > 50 else error['message'],
                        str(error.get('current_value', ''))[:30]
                    ])
                
                error_table = Table(error_data, colWidths=[0.8*inch, 1.2*inch, 3*inch, 1.5*inch])
                error_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.red),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.Color(1, 0.9, 0.9)),
                ]))
                story.append(error_table)
                story.append(Spacer(1, 20))
            
            # Warnings table
            if warnings:
                story.append(Paragraph("Avisos", self.styles['Heading3']))
                warning_data = [['Linha', 'Campo', 'Problema', 'Valor Atual']]
                
                for warning in warnings[:20]:  # Limit to 20 for PDF
                    warning_data.append([
                        str(warning['row']),
                        warning['field'],
                        warning['message'][:50] + '...' if len(warning['message']) > 50 else warning['message'],
                        str(warning.get('current_value', ''))[:30]
                    ])
                
                warning_table = Table(warning_data, colWidths=[0.8*inch, 1.2*inch, 3*inch, 1.5*inch])
                warning_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.orange),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.Color(1, 0.95, 0.85)),
                ]))
                story.append(warning_table)
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        return output.read()
    
    def get_content_type(self, format: str) -> str:
        """Get the appropriate content type for the format."""
        format = format.lower()
        content_types = {
            'json': 'application/json',
            'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'pdf': 'application/pdf'
        }
        return content_types.get(format, 'application/octet-stream')
    
    def get_file_extension(self, format: str) -> str:
        """Get the appropriate file extension for the format."""
        format = format.lower()
        extensions = {
            'json': 'json',
            'excel': 'xlsx',
            'xlsx': 'xlsx',
            'pdf': 'pdf'
        }
        return extensions.get(format, 'bin')